# scripts/build_election_playbook_data.py
#
# One-time / re-runnable importer: parses the "Oct 2026 Election Playbook" Excel
# workbook into data/election_playbook_2026.csv — one row per ticker (combo
# rows like "SPY / QQQ" or "SPXL / SPXS" are split into individual tickers).
#
# Re-run this manually if the user hands over an updated Excel workbook:
#   python scripts/build_election_playbook_data.py "path\to\New_Playbook.xlsx"

import re
import sys
import os
import openpyxl
import pandas as pd

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
OUT_CSV = os.path.join(DATA_DIR, "election_playbook_2026.csv")

# Leveraged / inverse ETFs called out by name in the "LEVERAGED / INVERSE ETF
# WARNING" section of the Sequencing & Legend sheet — every one resets daily.
LEVERAGED_TICKERS = {
    "CURE", "RXL", "FAS", "FAZ", "ERX", "ERY", "GUSH", "DRIP", "NRGU", "DFEN",
    "TNA", "TZA", "UWM", "SOXL", "SOXS", "NVDL", "NVDU", "SPXL", "SPXS",
    "UPRO", "SPXU", "TQQQ", "SQQQ",
}
ETN_TICKERS = {"NRGU"}  # issuer credit risk, not a true ETF

# Combo rows in the sheet that bundle 2 tickers into one row (bull/bear pair
# or a same-side pair). Keyed by the exact "Ticker & Type" cell text.
# Each entry: {ticker: (bull_bear, buy_on_dip_override_or_None)}
COMBO_ROWS = {
    "GUSH / DRIP (2x E&P bull/bear)": {
        "GUSH": ("Bull", "Y", "Tactical dip-buy"),
        "DRIP": ("Bear", "N", "Hedge-short (tactical)"),
    },
    "SOXX / SMH (ETFs)": {
        "SOXX": ("", "Y", None),
        "SMH":  ("", "Y", None),
    },
    "NVDL / NVDU (2x NVDA)": {
        "NVDL": ("Bull", "Y", None),
        "NVDU": ("Bull", "Y", None),
    },
    "SPY / QQQ (ETFs)": {
        "SPY": ("", "Partial", None),
        "QQQ": ("", "Partial", None),
    },
    "SPXL / SPXS (3x S&P)": {
        "SPXL": ("Bull", "Y", "Tactical dip-buy"),
        "SPXS": ("Bear", "N", "Hedge-short (tactical)"),
    },
    "UPRO / SPXU (3x S&P)": {
        "UPRO": ("Bull", "Y", "Tactical dip-buy"),
        "SPXU": ("Bear", "N", "Hedge-short (tactical)"),
    },
    "TQQQ / SQQQ (3x Nasdaq)": {
        "TQQQ": ("Bull", "Y", "Tactical dip-buy"),
        "SQQQ": ("Bear", "N", "Hedge-short (tactical)"),
    },
}

# Normalize the sheet's free-text Outlook into the 4-color legend key.
# ("Down (extended)" folds into "Down hard"; "Hedge" folds into "Volatile" —
# per user's choice to keep exactly the 4 legend colors.)
def normalize_outlook(raw: str) -> str:
    r = raw.lower()
    if "down hard" in r or "extended" in r:
        return "Down hard"
    if "down mild" in r:
        return "Down mild"
    if "hedge" in r or "volatile" in r:
        return "Volatile"
    if "resilient" in r or "election-agnostic" in r:
        return "Resilient"
    return "Volatile"


def election_beta(raw: str) -> str:
    """Extract the leading HIGH/MED/LOW beta token from the Election sensitivity cell."""
    m = re.match(r"\s*([A-Z][A-Z\-\s]*[A-Z])(?:\s*[–\-(]|\s*$)", raw)
    if m:
        return m.group(1).strip()
    return raw.strip()[:12]


def strategy_bucket(play_type: str) -> str:
    p = play_type.lower()
    if "avoid" in p:
        return "AVOID"
    if "hedge" in p:
        return "HEDGE"
    if "buy dip" in p and "csp" in p:
        return "DIP_OR_CSP"
    if "csp" in p:
        return "SELL_CSP"
    if "core" in p:
        return "CORE_HOLD"
    if "tactical" in p:
        return "TACTICAL_LEV"
    return "BUY_DIP"


def parse_instrument(ticker_type: str) -> str:
    m = re.search(r"\(([^)]*)\)", ticker_type)
    desc = (m.group(1) if m else "").lower()
    if "etn" in desc:
        return "ETN"
    if "etf" in desc:
        return "ETF"
    return "Stock"


def main(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Oct 2026 Playbook"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows_out = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        if not any(vals):
            continue
        (sector, ticker_type, outlook, why_drop, bounce_driver,
         election_sens, play_type, buy_on_dip, horizon, notes) = (vals + [None] * 10)[:10]

        ticker_type = str(ticker_type or "").strip()
        instrument = parse_instrument(ticker_type)
        outlook_cat = normalize_outlook(str(outlook or ""))
        el_beta = election_beta(str(election_sens or ""))

        if ticker_type in COMBO_ROWS:
            for tk, (bull_bear, dip_override, play_override) in COMBO_ROWS[ticker_type].items():
                pt = play_override or str(play_type or "")
                rows_out.append({
                    "Sector": sector, "Ticker": tk, "Instrument": instrument,
                    "Leveraged": "Y" if tk in LEVERAGED_TICKERS else "N",
                    "ETN": "Y" if tk in ETN_TICKERS else "N",
                    "BullBear": bull_bear,
                    "Outlook_Raw": outlook, "Outlook_Category": outlook_cat,
                    "Why_Drop": why_drop, "Bounce_Driver": bounce_driver,
                    "Election_Sensitivity": election_sens, "Election_Beta": el_beta,
                    "Play_Type": pt, "Strategy": strategy_bucket(pt),
                    "Buy_Only_On_Dips": dip_override,
                    "Horizon_Conviction": horizon, "Notes": notes,
                })
        else:
            m = re.match(r"^([A-Z]+)", ticker_type)
            ticker = m.group(1) if m else ticker_type.split()[0]
            rows_out.append({
                "Sector": sector, "Ticker": ticker, "Instrument": instrument,
                "Leveraged": "Y" if ticker in LEVERAGED_TICKERS else "N",
                "ETN": "Y" if ticker in ETN_TICKERS else "N",
                "BullBear": "",
                "Outlook_Raw": outlook, "Outlook_Category": outlook_cat,
                "Why_Drop": why_drop, "Bounce_Driver": bounce_driver,
                "Election_Sensitivity": election_sens, "Election_Beta": el_beta,
                "Play_Type": play_type, "Strategy": strategy_bucket(str(play_type or "")),
                "Buy_Only_On_Dips": buy_on_dip,
                "Horizon_Conviction": horizon, "Notes": notes,
            })

    df = pd.DataFrame(rows_out)
    os.makedirs(DATA_DIR, exist_ok=True)
    df.to_csv(OUT_CSV, index=False, encoding="utf-8")
    print(f"Wrote {len(df)} tickers -> {OUT_CSV}")
    dupes = df[df.duplicated("Ticker", keep=False)]
    if not dupes.empty:
        print("NOTE: duplicate tickers across sectors (kept, may need review):")
        print(dupes[["Sector", "Ticker"]].to_string(index=False))


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\llcai\Downloads\Oct_2026_Election_Playbook.xlsx"
    main(path)
